#!/usr/bin/env python3
"""
Yurtiçi Kargo - CrowdStrike Identity Protection Weekly Report
GraphQL tabanlı - tam veri
"""

import json
from datetime import datetime, timezone
from falconpy import IdentityProtection
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CLIENT_ID  = "a6df49f452c144ecaf630aae61b1d571"
SECRET     = "1JWgqADk6idMC4hpctzK7sxfXlTo3S02rE8OQb59"
BASE_URL   = "https://api.eu-1.crowdstrike.com"
OUTPUT_DIR = "/home/infosec/.openclaw/workspace/customers/reports/yurtici"

FONT="Helvetica"; ORANGE="F47920"; NAVY="1A2644"; WHITE="FFFFFF"
LIGHT_ORANGE="FDE8D5"; LIGHT_NAVY="D6DCE8"; GRAY="F5F5F5"
GREEN_BG="E2EFDA"; GREEN_FG="375623"; RED_BG="FFE0E0"; RED_FG="C00000"
YELLOW_BG="FFF2CC"; YELLOW_FG="7F6000"; PURPLE_BG="EAD1DC"; PURPLE_FG="4A235A"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(color="000000", bold=False, size=10, italic=False):
    return Font(color=color, bold=bold, size=size, name=FONT, italic=italic)
def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width
def fmt_date(d):
    if not d: return "—"
    try: return d[:10].replace("-","/") + " " + d[11:16]
    except: return str(d)[:16]

def page_all(idp, args, entity_fields):
    """Tüm sayfaları çek"""
    results = []
    after = None
    while True:
        after_str = f', after: "{after}"' if after else ''
        q = f'{{ entities({args}, first: 100{after_str}) {{ nodes {{ {entity_fields} }} pageInfo {{ hasNextPage endCursor }} }} }}'
        r = idp.graphql(query=q)
        if r['status_code'] != 200: break
        nodes = r['body']['data']['entities']['nodes']
        results.extend(nodes)
        pi = r['body']['data']['entities']['pageInfo']
        if not pi.get('hasNextPage'): break
        after = pi.get('endCursor')
    return results

def header_row(ws, row, headers, bg=None):
    bg = bg or ORANGE
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.fill = fill(bg); c.font = fnt(WHITE, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 22

def stat_card(ws, row, col, title, value, bg, fg):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    c1 = ws.cell(row=row, column=col, value=title)
    c1.fill = fill(bg); c1.font = fnt(fg, bold=True, size=9)
    c1.alignment = Alignment(horizontal="center", vertical="center"); c1.border = thin_border()
    c2 = ws.cell(row=row+1, column=col, value=value)
    c2.fill = fill(bg); c2.font = Font(color=fg, bold=True, size=22, name=FONT)
    c2.alignment = Alignment(horizontal="center", vertical="center"); c2.border = thin_border()
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row+1].height = 36

USER_FIELDS = """
... on UserEntity {
  primaryDisplayName
  secondaryDisplayName
  entityId
  riskScore
  riskScoreSeverity
  stale
  inactive
  shared
  watched
  hasADDomainAdminRole
  mostRecentActivity
  linkedAccountIds
  roles { type }
  riskFactors { type severity }
  accounts {
    ... on ActiveDirectoryAccountDescriptor {
      domain
      samAccountName
      enabled
      department
      ou
      title
      lockoutTime
      passwordAttributes { aged lastChange mayExpire exposed strength }
    }
  }
}
"""

def get_user_flags(u):
    """Kullanıcının flag'larını hesapla"""
    accs = u.get('accounts') or []
    risk_factors = [rf.get('type','') for rf in (u.get('riskFactors') or [])]
    
    pwd_never_expires = any(
        not (acc.get('passwordAttributes') or {}).get('mayExpire', True)
        for acc in accs
    )
    pwd_aged = any((acc.get('passwordAttributes') or {}).get('aged') for acc in accs)
    pwd_exposed = any((acc.get('passwordAttributes') or {}).get('exposed') for acc in accs)
    pwd_weak = any(
        (acc.get('passwordAttributes') or {}).get('strength') == 'Weak'
        for acc in accs
    )
    locked = any(acc.get('lockoutTime') for acc in accs)
    linked = bool(u.get('linkedAccountIds'))
    
    # Risk factor bazlı flag'lar
    stealthy = any('STEALTHY' in rf.upper() for rf in risk_factors)
    non_domain = any('NON_DOMAIN' in rf.upper() or 'UNMANAGED' in rf.upper() for rf in risk_factors)
    duplicate_pwd = any('DUPLICATE' in rf.upper() for rf in risk_factors)
    honeytoken = any('HONEY' in rf.upper() for rf in risk_factors)
    
    severity = (u.get('riskScoreSeverity') or '').upper()
    high_risk = severity in ('HIGH', 'CRITICAL')
    medium_risk = severity == 'MEDIUM'
    low_risk = severity == 'LOW'
    
    return {
        'Privileged':            u.get('hasADDomainAdminRole', False),
        'High Risk':             high_risk,
        'Medium Risk':           medium_risk,
        'Low Risk':              low_risk,
        'Stale':                 u.get('stale', False),
        'Inactive':              u.get('inactive', False),
        'Shared':                u.get('shared', False),
        'Watched':               u.get('watched', False),
        'Stealthy':              stealthy,
        'Non-Domain Joined':     non_domain,
        'Pwd Never Expires':     pwd_never_expires,
        'Compromised Pwd':       pwd_weak or pwd_exposed,
        'GPO Exposed Pwd':       pwd_exposed,
        'Aged Password':         pwd_aged,
        'Account Locked':        locked,
        'Linked Accounts':       linked,
        'Duplicate Password':    duplicate_pwd,
        'Honeytoken':            honeytoken,
    }

def main():
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%Y-%m-%d")
    today_display = now.strftime("%d/%m/%Y")

    print(f"[{date_str}] Yurtiçi Kargo Identity Protection raporu...")

    idp = IdentityProtection(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)

    # ── VERİ ÇEK ────────────────────────────────────────────
    print("  Kullanıcı verileri çekiliyor...")

    def count_entities(args):
        total = 0
        after = None
        while True:
            after_str = f', after: "{after}"' if after else ''
            q = f'{{ entities({args}, first: 100{after_str}) {{ nodes {{ entityId }} pageInfo {{ hasNextPage endCursor }} }} }}'
            r = idp.graphql(query=q)
            if r['status_code'] != 200: break
            total += len(r['body']['data']['entities']['nodes'])
            pi = r['body']['data']['entities']['pageInfo']
            if not pi.get('hasNextPage'): break
            after = pi.get('endCursor')
        return total

    total_users      = count_entities("types: [USER]")
    stale_users      = count_entities("types: [USER], stale: true")
    inactive_users   = count_entities("types: [USER], inactive: true")
    total_endpoints  = count_entities("types: [ENDPOINT]")
    high_risk_users       = count_entities("types: [USER], minRiskScoreSeverity: HIGH")
    pwd_never_expires     = count_entities("types: [USER], hasNeverExpiringPassword: true")
    compromised_pwd       = count_entities("types: [USER], hasWeakPassword: true")
    exposed_pwd           = count_entities("types: [USER], hasExposedPassword: true")
    shared_users          = count_entities("types: [USER], shared: true")
    watched_users         = count_entities("types: [USER], watched: true")
    marked_users          = count_entities("types: [USER], marked: true")
    locked_users          = count_entities("types: [USER], accountLocked: true")
    aged_pwd_users        = count_entities("types: [USER], hasAgedPassword: true")

    print(f"  Users: {total_users} | Stale: {stale_users} | Inactive: {inactive_users}")
    print(f"  High Risk: {high_risk_users} | Pwd Never Expires: {pwd_never_expires} | Compromised: {compromised_pwd}")
    print(f"  Exposed Pwd: {exposed_pwd} | Shared: {shared_users} | Locked: {locked_users} | Aged Pwd: {aged_pwd_users}")

    # Privileged roles
    priv_roles = {
        "Domain Admins":    "DomainAdminsRole",
        "Enterprise Admins":"EnterpriseAdminsRole",
        "Schema Admins":    "SchemaAdminsRole",
        "Administrators":   "AdministratorsRole",
        "Local Admins":     "LocalAdminRole",
    }
    priv_counts = {}
    for label, role in priv_roles.items():
        priv_counts[label] = count_entities(f"types: [USER], roles: [{role}]")
    print(f"  Privileged: {priv_counts}")

    # Tüm kullanıcıları çek (max 3000)
    print("  Tüm kullanıcılar çekiliyor...")
    all_users = []
    after = None
    while len(all_users) < 3000:
        after_str = f', after: "{after}"' if after else ''
        q = f'{{ entities(types: [USER], first: 100{after_str}) {{ nodes {{ {USER_FIELDS} }} pageInfo {{ hasNextPage endCursor }} }} }}'
        r = idp.graphql(query=q)
        if r['status_code'] != 200: break
        nodes = r['body']['data']['entities']['nodes']
        all_users.extend(nodes)
        pi = r['body']['data']['entities']['pageInfo']
        if not pi.get('hasNextPage'): break
        after = pi.get('endCursor')
    print(f"  {len(all_users)} kullanıcı çekildi")

    # Privileged users (ayrı çek - tam liste)
    print("  Privileged users çekiliyor...")
    priv_users = {}
    for label, role in priv_roles.items():
        priv_users[label] = page_all(idp, f"types: [USER], roles: [{role}]", USER_FIELDS)
    total_priv = sum(len(v) for v in priv_users.values())
    print(f"  Toplam privileged: {total_priv}")

    # Stale users
    stale_list = page_all(idp, "types: [USER], stale: true", USER_FIELDS)
    print(f"  Stale users: {len(stale_list)}")

    # Risk score hesapla
    risk_score = 0.0
    risk_factors = []
    stale_pct = (stale_users / total_users * 100) if total_users else 0
    inactive_pct = (inactive_users / total_users * 100) if total_users else 0
    da_count = priv_counts.get("Domain Admins", 0)
    ea_count = priv_counts.get("Enterprise Admins", 0)

    # Şifre sorunları
    pwd_aged = sum(1 for u in all_users for acc in (u.get('accounts') or []) if (acc.get('passwordAttributes') or {}).get('aged'))
    pwd_exposed = sum(1 for u in all_users for acc in (u.get('accounts') or []) if (acc.get('passwordAttributes') or {}).get('exposed'))

    if stale_pct > 10: risk_score += 2; risk_factors.append(("High Stale Account Ratio", f"{stale_users:,} accounts ({stale_pct:.1f}%)", "HIGH"))
    elif stale_pct > 5: risk_score += 1; risk_factors.append(("Stale Accounts", f"{stale_users:,} accounts ({stale_pct:.1f}%)", "MEDIUM"))

    if inactive_pct > 30: risk_score += 2; risk_factors.append(("High Inactive Account Ratio", f"{inactive_users:,} accounts ({inactive_pct:.1f}%)", "HIGH"))
    elif inactive_pct > 20: risk_score += 1; risk_factors.append(("Inactive Accounts", f"{inactive_users:,} accounts ({inactive_pct:.1f}%)", "MEDIUM"))

    if da_count > 10: risk_score += 2; risk_factors.append(("Too Many Domain Admins", f"{da_count} accounts", "HIGH"))
    elif da_count > 5: risk_score += 1; risk_factors.append(("Domain Admin Count", f"{da_count} accounts", "MEDIUM"))

    if ea_count > 3: risk_score += 1.5; risk_factors.append(("Enterprise Admins", f"{ea_count} accounts", "HIGH"))

    if pwd_exposed > 0: risk_score += 2; risk_factors.append(("Exposed Passwords Detected", f"{pwd_exposed} accounts", "CRITICAL"))
    if pwd_aged > 100: risk_score += 1; risk_factors.append(("Aged Passwords", f"{pwd_aged} accounts", "MEDIUM"))

    risk_score = min(round(risk_score * 1.2, 1), 10)

    # ── WORKBOOK ────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ── COVER ───────────────────────────────────────────────
    ws_cover = wb.create_sheet("Cover")
    ws_cover.sheet_properties.tabColor = NAVY
    ws_cover.sheet_view.showGridLines = False
    for row in range(1, 50):
        for col in range(1, 15):
            ws_cover.cell(row=row, column=col).fill = fill(WHITE)
    for col in range(1, 15):
        for r in range(1, 6):
            ws_cover.cell(row=r, column=col).fill = fill(NAVY)
    ws_cover.merge_cells("B2:N4")
    ws_cover["B2"] = "CrowdStrike Falcon"
    ws_cover["B2"].font = Font(color=WHITE, bold=True, size=28, name=FONT)
    ws_cover["B2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 15): ws_cover.cell(row=6, column=col).fill = fill(ORANGE)
    ws_cover.row_dimensions[6].height = 6
    ws_cover.merge_cells("B8:N9")
    ws_cover["B8"] = "Yurtiçi Kargo"
    ws_cover["B8"].font = Font(color=NAVY, bold=True, size=32, name=FONT)
    ws_cover["B8"].alignment = Alignment(horizontal="left", vertical="center")
    ws_cover.merge_cells("B10:N11")
    ws_cover["B10"] = "Identity Protection Health Check Report"
    ws_cover["B10"].font = Font(color=ORANGE, bold=True, size=20, name=FONT)
    ws_cover["B10"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 14): ws_cover.cell(row=12, column=col).fill = fill(ORANGE)
    ws_cover.row_dimensions[12].height = 3
    for i, (label, value) in enumerate([
        ("Report Date", today_display), ("Prepared by", "Pure7"),
        ("Customer", "Yurtiçi Kargo"), ("Product", "CrowdStrike Falcon Identity Protection"),
        ("Data Source", "GraphQL API"), ("Version", "2.0"),
    ]):
        r = 14 + i * 2
        ws_cover.merge_cells(f"B{r}:D{r}"); ws_cover[f"B{r}"] = label
        ws_cover[f"B{r}"].font = fnt(NAVY, bold=True, size=11)
        ws_cover[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws_cover.merge_cells(f"E{r}:N{r}"); ws_cover[f"E{r}"] = value
        ws_cover[f"E{r}"].font = fnt("333333", size=11)
        ws_cover[f"E{r}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 15):
        for r in range(45, 50): ws_cover.cell(row=r, column=col).fill = fill(NAVY)
    ws_cover.merge_cells("B46:N48")
    ws_cover["B46"] = "Pure7 InfoSec Reports  |  Confidential"
    ws_cover["B46"].font = Font(color=WHITE, size=10, name=FONT)
    ws_cover["B46"].alignment = Alignment(horizontal="center", vertical="center")
    for r in range(1, 50): ws_cover.row_dimensions[r].height = 20
    for r in range(2, 5): ws_cover.row_dimensions[r].height = 28
    ws_cover.row_dimensions[8].height = 40; ws_cover.row_dimensions[10].height = 30
    set_col_width(ws_cover, 1, 3)
    for c in range(2, 15): set_col_width(ws_cover, c, 12)

    # ── SUMMARY / DASHBOARD ─────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = ORANGE
    ws_sum.sheet_view.showGridLines = False
    ws_sum.merge_cells("A1:L1")
    ws_sum["A1"] = "Yurtiçi Kargo – Identity Protection Dashboard"
    ws_sum["A1"].fill = fill(NAVY); ws_sum["A1"].font = fnt(WHITE, bold=True, size=16)
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 40
    ws_sum.merge_cells("A2:L2")
    ws_sum["A2"] = f"Report Date: {today_display}  |  Data Source: CrowdStrike GraphQL API"
    ws_sum["A2"].fill = fill(ORANGE); ws_sum["A2"].font = fnt(WHITE, bold=True, size=10)
    ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    # Risk Score
    risk_color = RED_FG if risk_score >= 7 else (YELLOW_FG if risk_score >= 4 else GREEN_FG)
    risk_bg = RED_BG if risk_score >= 7 else (YELLOW_BG if risk_score >= 4 else GREEN_BG)
    ws_sum.merge_cells("A4:D4")
    ws_sum["A4"] = "  IDENTITY RISK SCORE"
    ws_sum["A4"].fill = fill(LIGHT_NAVY); ws_sum["A4"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[4].height = 22
    ws_sum.merge_cells("A5:D6")
    ws_sum["A5"] = f"{risk_score} / 10"
    ws_sum["A5"].fill = fill(risk_bg)
    ws_sum["A5"].font = Font(color=risk_color, bold=True, size=36, name=FONT)
    ws_sum["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[5].height = 25; ws_sum.row_dimensions[6].height = 25

    # User Overview
    ws_sum.merge_cells("E4:L4")
    ws_sum["E4"] = "  USER OVERVIEW"
    ws_sum["E4"].fill = fill(LIGHT_NAVY); ws_sum["E4"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["E4"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[4].height = 22
    stat_card(ws_sum, 5, 5, "Total Users",    total_users,    LIGHT_NAVY,  NAVY)
    stat_card(ws_sum, 5, 7, "Stale",          stale_users,    RED_BG,      RED_FG)
    stat_card(ws_sum, 5, 9, "Inactive (30d)", inactive_users, YELLOW_BG,   YELLOW_FG)
    stat_card(ws_sum, 5, 11,"Total Endpoints",total_endpoints, LIGHT_NAVY, NAVY)

    # Privileged Overview
    ws_sum.merge_cells("A8:L8")
    ws_sum["A8"] = "  PRIVILEGED ACCOUNTS"
    ws_sum["A8"].fill = fill(LIGHT_NAVY); ws_sum["A8"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A8"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[8].height = 22
    stat_card(ws_sum, 9, 1, "Domain Admins",    priv_counts.get("Domain Admins",0),    RED_BG,    RED_FG)
    stat_card(ws_sum, 9, 3, "Enterprise Admins",priv_counts.get("Enterprise Admins",0),RED_BG,    RED_FG)
    stat_card(ws_sum, 9, 5, "Schema Admins",    priv_counts.get("Schema Admins",0),    YELLOW_BG, YELLOW_FG)
    stat_card(ws_sum, 9, 7, "Administrators",   priv_counts.get("Administrators",0),   YELLOW_BG, YELLOW_FG)
    stat_card(ws_sum, 9, 9, "Local Admins",     priv_counts.get("Local Admins",0),     LIGHT_NAVY,NAVY)

    # Password Health
    ws_sum.merge_cells("A12:L12")
    ws_sum["A12"] = "  PASSWORD HEALTH & USER RISK"
    ws_sum["A12"].fill = fill(LIGHT_NAVY); ws_sum["A12"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A12"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[12].height = 22
    stat_card(ws_sum, 13, 1,  "🔴 High Risk",          high_risk_users,    RED_BG,    RED_FG)
    stat_card(ws_sum, 13, 3,  "Pwd Never Expires",     pwd_never_expires,  RED_BG if pwd_never_expires>100 else YELLOW_BG, RED_FG if pwd_never_expires>100 else YELLOW_FG)
    stat_card(ws_sum, 13, 5,  "Compromised Pwd",       compromised_pwd,    RED_BG if compromised_pwd>0 else GREEN_BG, RED_FG if compromised_pwd>0 else GREEN_FG)
    stat_card(ws_sum, 13, 7,  "GPO Exposed Pwd",       exposed_pwd,        RED_BG if exposed_pwd>0 else GREEN_BG, RED_FG if exposed_pwd>0 else GREEN_FG)
    stat_card(ws_sum, 13, 9,  "Aged Passwords",        aged_pwd_users,     RED_BG if aged_pwd_users>100 else YELLOW_BG, RED_FG if aged_pwd_users>100 else YELLOW_FG)
    stat_card(ws_sum, 13, 11, "🔒 Account Locked",     locked_users,       RED_BG if locked_users>0 else GREEN_BG, RED_FG if locked_users>0 else GREEN_FG)

    # User Flags
    ws_sum.merge_cells("A16:L16")
    ws_sum["A16"] = "  USER FLAGS"
    ws_sum["A16"].fill = fill(LIGHT_NAVY); ws_sum["A16"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A16"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[16].height = 22
    stat_card(ws_sum, 17, 1, "Shared Accounts",  shared_users,  YELLOW_BG if shared_users>0 else GREEN_BG, YELLOW_FG if shared_users>0 else GREEN_FG)
    stat_card(ws_sum, 17, 3, "Watched Users",    watched_users, YELLOW_BG if watched_users>0 else GREEN_BG, YELLOW_FG if watched_users>0 else GREEN_FG)
    stat_card(ws_sum, 17, 5, "Marked Users",     marked_users,  YELLOW_BG if marked_users>0 else GREEN_BG, YELLOW_FG if marked_users>0 else GREEN_FG)

    # Risk Factors tablosu
    ws_sum.merge_cells("A20:L20")
    ws_sum["A20"] = "  RISK FACTORS"
    ws_sum["A20"].fill = fill(LIGHT_NAVY); ws_sum["A20"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A20"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[20].height = 22

    for col_i, h in enumerate(["Risk Factor","Value","Severity"],1):
        c = ws_sum.cell(row=21, column=col_i*2-1, value=h)
        ws_sum.merge_cells(start_row=21, start_column=col_i*2-1, end_row=21, end_column=col_i*2)
        c.fill = fill(ORANGE); c.font = fnt(WHITE, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
    ws_sum.row_dimensions[21].height = 18

    sev_colors = {"CRITICAL":(RED_FG,RED_BG),"HIGH":(RED_FG,RED_BG),"MEDIUM":(YELLOW_FG,YELLOW_BG),"LOW":(GREEN_FG,GREEN_BG)}
    if not risk_factors:
        ws_sum.merge_cells("A22:F22")
        ws_sum["A22"] = "✅ No significant risk factors detected"
        ws_sum["A22"].fill = fill(GREEN_BG); ws_sum["A22"].font = fnt(GREEN_FG, bold=True, size=10)
        ws_sum["A22"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        for i, (factor, value, severity) in enumerate(risk_factors):
            r = 22 + i
            fg, bg = sev_colors.get(severity, ("000000", WHITE))
            ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws_sum.cell(row=r, column=1, value=factor).fill = fill(bg)
            ws_sum.cell(row=r, column=1).font = fnt(size=9); ws_sum.cell(row=r, column=1).border = thin_border()
            ws_sum.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            ws_sum.cell(row=r, column=3, value=value).fill = fill(bg)
            ws_sum.cell(row=r, column=3).font = fnt(size=9); ws_sum.cell(row=r, column=3).border = thin_border()
            ws_sum.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
            c = ws_sum.cell(row=r, column=5, value=severity)
            c.fill = fill(bg); c.font = fnt(fg, bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
            ws_sum.row_dimensions[r].height = 18
    for c in range(1, 13): set_col_width(ws_sum, c, 10)

    # ── PRIVILEGED ACCOUNTS ─────────────────────────────────
    print("  Privileged Accounts sheet...")
    ws_priv = wb.create_sheet("Privileged Accounts")
    ws_priv.sheet_properties.tabColor = RED_FG

    # Tüm privileged kullanıcıları birleştir (unique)
    seen = set()
    all_priv = []
    for role_label, users in priv_users.items():
        for u in users:
            uid = u.get('entityId','')
            if uid not in seen:
                seen.add(uid)
                u['_role_label'] = role_label
                all_priv.append(u)
            else:
                for existing in all_priv:
                    if existing.get('entityId') == uid:
                        existing['_role_label'] = existing.get('_role_label','') + ', ' + role_label
                        break

    FLAG_COLS = [
        "Privileged","High Risk","Medium Risk","Stale","Inactive","Shared","Watched",
        "Stealthy","Non-Domain Joined","Pwd Never Expires","Compromised Pwd",
        "GPO Exposed Pwd","Aged Password","Account Locked","Linked Accounts",
        "Duplicate Password","Honeytoken"
    ]
    PRIV_H = ["Display Name","SAM Account","Domain","Department","Title",
              "Privileged Role(s)","Risk Score","Severity","Last Activity"] + FLAG_COLS

    ws_priv.merge_cells(f"A1:{get_column_letter(len(PRIV_H))}1")
    ws_priv["A1"] = f"Yurtiçi Kargo – Privileged Accounts  ({len(all_priv)} unique)"
    ws_priv["A1"].fill = fill(NAVY); ws_priv["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_priv["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_priv.row_dimensions[1].height = 35
    header_row(ws_priv, 2, PRIV_H)

    for idx, u in enumerate(sorted(all_priv, key=lambda x: x.get('riskScore',0) or 0, reverse=True), 1):
        row = idx + 2
        accs = u.get('accounts') or []
        acc = accs[0] if accs else {}
        risk = u.get('riskScore') or 0
        sev = u.get('riskScoreSeverity','—') or '—'
        bg = RED_BG if risk >= 0.7 else (YELLOW_BG if risk >= 0.4 else (WHITE if idx%2!=0 else GRAY))
        flags = get_user_flags(u)

        row_data = [
            u.get('primaryDisplayName','—'),
            acc.get('samAccountName','—'),
            acc.get('domain','—'),
            acc.get('department','—') or '—',
            acc.get('title','—') or '—',
            u.get('_role_label','—'),
            f"{risk:.2f}",
            sev,
            fmt_date(u.get('mostRecentActivity','')),
        ] + ["●" if flags.get(fc) else "" for fc in FLAG_COLS]

        for col_i, val in enumerate(row_data, 1):
            c = ws_priv.cell(row=row, column=col_i, value=val)
            c.fill = fill(bg); c.border = thin_border()
            is_flag = col_i > 9
            c.alignment = Alignment(horizontal="center" if is_flag or col_i in (7,8) else "left", vertical="center")
            if col_i == 7:
                c.font = fnt(RED_FG if risk>=0.7 else YELLOW_FG if risk>=0.4 else GREEN_FG, bold=True, size=9)
            elif is_flag and val == "●":
                c.font = fnt(RED_FG, bold=True, size=9)
            else:
                c.font = fnt(NAVY if col_i==1 else "000000", bold=(col_i==1), size=9)
        ws_priv.row_dimensions[row].height = 16

    ws_priv.freeze_panes = "A3"
    for col_i, w in enumerate([28,18,18,18,16,28,10,12,18]+[10]*len(FLAG_COLS), 1):
        set_col_width(ws_priv, col_i, w)

    # ── STALE ACCOUNTS ──────────────────────────────────────
    print("  Stale Accounts sheet...")
    ws_stale = wb.create_sheet("Stale Accounts")
    ws_stale.sheet_properties.tabColor = YELLOW_FG

    ws_stale.merge_cells(f"A1:{get_column_letter(8)}1")
    ws_stale["A1"] = f"Yurtiçi Kargo – Stale Accounts  ({len(stale_list)} accounts)"
    ws_stale["A1"].fill = fill(NAVY); ws_stale["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_stale["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_stale.row_dimensions[1].height = 35

    STALE_H = ["Display Name","Account","Domain","Department","Risk Score","Domain Admin","Pwd Aged","Last Activity"]
    header_row(ws_stale, 2, STALE_H)

    for idx, u in enumerate(sorted(stale_list, key=lambda x: x.get('riskScore',0) or 0, reverse=True), 1):
        row = idx + 2
        accs = u.get('accounts') or []
        acc = accs[0] if accs else {}
        pwd = acc.get('passwordAttributes') or {}
        risk = u.get('riskScore') or 0
        bg = YELLOW_BG if idx%2!=0 else WHITE

        row_data = [
            u.get('primaryDisplayName','—'),
            acc.get('samAccountName','—'),
            acc.get('domain','—'),
            acc.get('department','—') or '—',
            f"{risk:.2f}",
            "✅" if u.get('hasADDomainAdminRole') else "",
            "●" if pwd.get('aged') else "",
            fmt_date(u.get('mostRecentActivity','')),
        ]
        for col_i, val in enumerate(row_data, 1):
            c = ws_stale.cell(row=row, column=col_i, value=val)
            c.fill = fill(bg); c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if col_i in (5,6,7) else "left", vertical="center")
            c.font = fnt(NAVY if col_i==1 else "000000", bold=(col_i==1), size=9)
        ws_stale.row_dimensions[row].height = 16

    ws_stale.freeze_panes = "A3"
    for col_i, w in enumerate([28,18,18,20,10,12,10,18],1): set_col_width(ws_stale, col_i, w)

    # ── ALL USERS ───────────────────────────────────────────
    print("  All Users sheet...")
    ws_users = wb.create_sheet("All Users")
    ws_users.sheet_properties.tabColor = NAVY

    USERS_H = ["Display Name","SAM Account","Domain","Department","Title",
               "Risk Score","Severity","Last Activity"] + FLAG_COLS

    ws_users.merge_cells(f"A1:{get_column_letter(len(USERS_H))}1")
    ws_users["A1"] = f"Yurtiçi Kargo – User Inventory  ({len(all_users)} of {total_users})"
    ws_users["A1"].fill = fill(NAVY); ws_users["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_users["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_users.row_dimensions[1].height = 35
    header_row(ws_users, 2, USERS_H)

    for idx, u in enumerate(sorted(all_users, key=lambda x: x.get('riskScore',0) or 0, reverse=True), 1):
        row = idx + 2
        accs = u.get('accounts') or []
        acc = accs[0] if accs else {}
        risk = u.get('riskScore') or 0
        sev = u.get('riskScoreSeverity','—') or '—'
        bg = RED_BG if risk >= 0.7 else (YELLOW_BG if risk >= 0.4 else (WHITE if idx%2!=0 else GRAY))
        flags = get_user_flags(u)

        row_data = [
            u.get('primaryDisplayName','—'),
            acc.get('samAccountName','—'),
            acc.get('domain','—'),
            acc.get('department','—') or '—',
            acc.get('title','—') or '—',
            f"{risk:.2f}",
            sev,
            fmt_date(u.get('mostRecentActivity','')),
        ] + ["●" if flags.get(fc) else "" for fc in FLAG_COLS]

        for col_i, val in enumerate(row_data, 1):
            c = ws_users.cell(row=row, column=col_i, value=val)
            c.fill = fill(bg); c.border = thin_border()
            is_flag = col_i > 8
            c.alignment = Alignment(horizontal="center" if is_flag or col_i in (6,7) else "left", vertical="center")
            if col_i == 6:
                c.font = fnt(RED_FG if risk>=0.7 else YELLOW_FG if risk>=0.4 else GREEN_FG, bold=True, size=9)
            elif is_flag and val == "●":
                c.font = fnt(RED_FG, bold=True, size=9)
            else:
                c.font = fnt(NAVY if col_i==1 else "000000", bold=(col_i==1), size=9)
        ws_users.row_dimensions[row].height = 16

    ws_users.freeze_panes = "A3"
    for col_i, w in enumerate([28,18,18,20,16,10,12,18]+[10]*len(FLAG_COLS), 1):
        set_col_width(ws_users, col_i, w)

    # ── KAYDET ──────────────────────────────────────────────
    filename = f"Yurtici_Kargo_Identity_Protection_HealthCheck_{date_str}.xlsx"
    output_path = f"{OUTPUT_DIR}/{filename}"
    wb.save(output_path)
    print(f"✅ Rapor kaydedildi: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
