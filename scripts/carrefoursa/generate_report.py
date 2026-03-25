#!/usr/bin/env python3
"""
CarrefourSA CrowdStrike Weekly Health Check Report
Her Cuma 15:00 (Istanbul) otomatik çalışır.
"""

import json, ast
from datetime import datetime, timezone
from falconpy import (
    Hosts, HostGroup, PreventionPolicy, SensorUpdatePolicy, 
    FirewallPolicies, DeviceControlPolicies, ResponsePolicies, Discover,
    MLExclusions, IOAExclusions, SensorVisibilityExclusions
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, Outline

# ── CONFIG ──────────────────────────────────────────────────
CLIENT_ID  = "4c6024b3d8524613b5c37abe055140e5"
SECRET     = "O4E7rwYKC2cBsPletv0X5pU9kfbao6FI31uzNx8H"
BASE_URL   = "https://api.eu-1.crowdstrike.com"
OUTPUT_DIR = "/home/infosec/.openclaw/workspace/customers/reports/carrefoursa"

# ── STYLE ───────────────────────────────────────────────────
FONT       = "Helvetica"
ORANGE     = "F47920"
NAVY       = "1A2644"
WHITE      = "FFFFFF"
LIGHT_ORANGE = "FDE8D5"
LIGHT_NAVY = "D6DCE8"
GRAY       = "F5F5F5"
GREEN_BG   = "E2EFDA"
GREEN_FG   = "375623"
RED_BG     = "FFE0E0"
RED_FG     = "C00000"
YELLOW_BG  = "FFF2CC"
YELLOW_FG  = "7F6000"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(color="000000", bold=False, size=10, italic=False):
    return Font(color=color, bold=bold, size=size, name=FONT, italic=italic)
def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width
def fmt_date(d):
    if not d: return "—"
    return d[:10].replace("-","/") + " " + d[11:16] if len(d) > 10 else d
def set_outline(ws):
    ws.sheet_properties = WorksheetProperties()
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)
def hide_row(ws, r, level):
    ws.row_dimensions[r].outline_level = level
    ws.row_dimensions[r].hidden = True

DESCRIPTIONS = {
    "Notify End Users": "Sends a notification to the end user when a threat is detected.",
    "Sensor Tampering Protection": "Prevents the sensor from being disabled or manipulated.",
    "Additional User Mode Data": "Collects additional telemetry from user-mode applications.",
    "Interpreter-Only": "Monitors interpreter-based activities (PowerShell, Python, etc.).",
    "Engine (Full Visibility)": "Collects comprehensive data from all application engines.",
    "Script-Based Execution Monitoring": "Monitors and detects script-based attacks.",
    "HTTP Detections": "Detects suspicious activities over HTTP traffic.",
    "Redact HTTP Detection Details": "Redacts sensitive information in HTTP detections.",
    "Hardware-Enhanced Exploit Detection": "Detects exploit attempts using hardware-based security features.",
    "Enhanced Exploitation Visibility": "Provides additional telemetry for exploit attempts.",
    "Extended User Mode Data": "Collects extended data from user-mode processes.",
    "Enhanced DLL Load Visibility": "Monitors DLL loading activities in detail.",
    "WSL2 Visibility": "Monitors activities inside Windows Subsystem for Linux 2.",
    "Memory Scanning": "Scans memory for suspicious code and patterns.",
    "Scan with CPU": "Uses CPU security features to perform memory scanning.",
    "BIOS Deep Visibility": "Deeply monitors BIOS/UEFI firmware activities.",
    "Cloud Anti-malware": "Detects and prevents malware using cloud-based ML.",
    "Adware & PUP": "Detects adware and potentially unwanted programs.",
    "Sensor Anti-malware": "Detects malware offline using the endpoint ML engine.",
    "Enhanced ML for larger files": "Applies enhanced ML analysis for larger files.",
    "Sensor Anti-malware for End-User Initiated Scans": "Uses sensor ML engine for user-initiated scans.",
    "Cloud Anti-malware for End-User Initiated Scans": "Uses cloud ML engine for user-initiated scans.",
    "Cloud PUP/Adware for End-User Initiated Scans": "Applies cloud analysis for PUP/Adware in user scans.",
    "USB Insertion Triggered Scan": "Triggers automatic scan when a USB device is inserted.",
    "Detect on Write": "Detects malicious content when a file is written.",
    "Quarantine on Write": "Quarantines the file automatically when a threat is detected on write.",
    "On Write Script File Visibility": "Provides visibility when script files are written.",
    "Quarantine & Security Center Registration": "Integrates quarantine with Windows Security Center.",
    "Quarantine on Removable Media": "Applies quarantine for threats on removable media.",
    "Cloud Anti-malware For Microsoft Office Files": "Detects malicious macros in Office files using cloud ML.",
    "Microsoft Office File Malicious Macro Removal": "Removes detected malicious macros from Office files.",
    "Custom Blocking": "Blocks executions based on custom-defined IOCs or rules.",
    "Suspicious Processes": "Detects and blocks processes with suspicious behavior.",
    "Suspicious Registry Operations": "Detects and blocks malicious registry changes.",
    "Boot Configuration Database Protection": "Blocks unauthorized changes to the boot configuration database.",
    "File System Containment": "Restricts file system changes on a compromised system.",
    "Suspicious Scripts and Commands": "Blocks malicious command-line and script executions.",
    "Intelligence-Sourced Threats": "Blocks threats based on threat intelligence IOCs.",
    "Driver Load Prevention": "Prevents unsigned or malicious drivers from loading.",
    "Vulnerable Driver Protection": "Prevents exploitation of known vulnerable drivers.",
    "Force ASLR": "Enforces Address Space Layout Randomization on all processes.",
    "Force DEP": "Enforces Data Execution Prevention on all processes.",
    "Heap Spray Preallocation": "Pre-allocates memory regions to prevent heap spray attacks.",
    "NULL Page Allocation": "Allocates the NULL page to prevent NULL pointer dereference exploits.",
    "SEH Overwrite Protection": "Protects against SEH overwrite attacks.",
    "Backup Deletion": "Detects and blocks ransomware attempts to delete backup files.",
    "Cryptowall": "Detects and blocks Cryptowall ransomware behavior.",
    "File Encryption": "Blocks unauthorized bulk file encryption attempts.",
    "Locky": "Detects and blocks Locky ransomware behavior.",
    "File System Access": "Detects ransomware-specific bulk file system access patterns.",
    "Volume Shadow Copy - Audit": "Detects and reports shadow copy deletion attempts.",
    "Volume Shadow Copy - Protect": "Actively blocks shadow copy deletion attempts.",
    "Application Exploitation Activity": "Detects behaviors indicating application exploit attempts.",
    "Chopper Webshell": "Detects and blocks use of the Chopper web shell tool.",
    "Drive-by Download": "Detects automatic malware downloads via browser.",
    "Code Injection": "Blocks code injection attempts into process memory.",
    "JavaScript Execution Via Rundll32": "Blocks JavaScript execution via Rundll32.",
    'Windows Logon Bypass ("Sticky Keys")': "Blocks logon bypass attempts via Sticky Keys.",
    "Windows Logon Bypass (Sticky Keys)": "Blocks logon bypass attempts via Sticky Keys.",
    "Credential Dumping": "Blocks credential theft attempts from sources like LSASS.",
    "Advanced Remediation": "Enables advanced remediation tools for automatic threat cleanup.",
    "Cloud-based Anomalous Process Execution": "Detects anomalous process execution behaviors using cloud analysis.",
    "Build": "Shows the build number in the sensor update configuration.",
    "Stage": "Specifies the deployment stage of the sensor update.",
    "Uninstall Protection": "Prevents unauthorized removal of the sensor.",
}

LEVEL_COLORS = {
    "AGGRESSIVE": ("FF0000", "FFE0E0"),
    "MODERATE":   ("C55A00", "FFE8CC"),
    "CAUTIOUS":   (YELLOW_FG, YELLOW_BG),
    "DISABLED":   ("808080", GRAY),
    "ON":         (GREEN_FG, GREEN_BG),
    "OFF":        ("808080", GRAY),
}

def level_style(val):
    return LEVEL_COLORS.get(str(val).upper().strip(), (NAVY, LIGHT_NAVY))

def parse_status(raw):
    if raw in ("ON", "OFF"):
        return "simple", raw, None
    try:
        if isinstance(raw, str) and "{" in raw:
            val = ast.literal_eval(raw)
            if isinstance(val, dict):
                return "mixed", val.get("detection","—").upper(), val.get("prevention","—").upper()
    except:
        pass
    return "simple", raw, None

def main():
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    now_istanbul = datetime.now(timezone.utc)
    date_str = now_istanbul.strftime("%Y-%m-%d")
    today_display = now_istanbul.strftime("%d/%m/%Y")
    report_month = now_istanbul.strftime("%B %Y")

    print(f"[{date_str}] CarrefourSA raporu oluşturuluyor...")

    # ── INITIALIZE FALCONPY CLIENTS ────────────────────────
    print("  FalconPy clients initializing...")
    hosts_client = Hosts(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    hg_client = HostGroup(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    prevent_client = PreventionPolicy(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    sensor_client = SensorUpdatePolicy(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    firewall_client = FirewallPolicies(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    devctrl_client = DeviceControlPolicies(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    response_client = ResponsePolicies(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    discover_client = Discover(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)

    # ── HOSTLARI ÇEK (FALCONPY) ────────────────────────────
    print("  Hostlar çekiliyor...")
    all_ids = []
    offset = 0
    while True:
        r = hosts_client.query_devices_by_filter(limit=500, offset=offset)
        if r['status_code'] != 200:
            print(f"  ⚠️  Error querying device IDs: {r.get('errors', [])}")
            break
        ids = r.get("body", {}).get("resources", [])
        if not ids: 
            break
        all_ids.extend(ids)
        offset += len(ids)
        if len(ids) < 500: 
            break

    hosts = []
    for i in range(0, len(all_ids), 100):
        batch = all_ids[i:i+100]
        r = hosts_client.get_device_details(ids=batch)
        if r['status_code'] == 200:
            hosts.extend(r.get("body", {}).get("resources", []))
    print(f"  {len(hosts)} host çekildi.")

    # ── GRUPLARI ÇEK (FALCONPY) ────────────────────────────
    r = hg_client.query_combined_host_groups(limit=500)
    group_list = []
    if r['status_code'] == 200:
        group_list = r.get("body", {}).get("resources", [])
    groups = {g["id"]: g["name"] for g in group_list}

    # ── POLİCY'LERİ ÇEK (FALCONPY) ─────────────────────────
    print("  Policy'ler çekiliyor...")
    policies = {}
    for ptype, client_obj in [("prevention", prevent_client), ("sensor_update", sensor_client)]:
        r = client_obj.query_combined_policies(limit=100)
        if r['status_code'] == 200:
            policies_list = r.get("body", {}).get("resources", [])
            for p in policies_list:
                policies[p["id"]] = {"name": p["name"], "type": ptype}

    # Policy detay verisi (grup bazında) - USING FALCONPY
    POLICY_CLIENTS = {
        "Prevention":     prevent_client,
        "Sensor Update":  sensor_client,
        "Firewall":       firewall_client,
        "Device Control": devctrl_client,
        "Response":       response_client,
    }
    
    policy_data = {}
    for g in group_list:
        gname = g.get("name"); gid = g.get("id")
        policy_data[gname] = {}
        for ptype, client_obj in POLICY_CLIENTS.items():
            filter_str = f"groups:['{gid}']"
            r = client_obj.query_combined_policies(filter=filter_str, limit=100)
            if r['status_code'] == 200 and r.get("body"):
                policy_data[gname][ptype] = []
                for p in r.get("body", {}).get("resources", []) or []:
                            pd = {"name": p.get("name"), "enabled": p.get("enabled"),
                                  "platform": p.get("platform_name",""), "settings": {}}
                            for cat in p.get("prevention_settings", []):
                                cat_name = cat.get("name","")
                                pd["settings"][cat_name] = []
                                for s in cat.get("settings", []):
                                    val = s.get("value", {})
                                    if isinstance(val, dict):
                                        en = val.get("enabled", None)
                                        if en is True: status = "ON"
                                        elif en is False: status = "OFF"
                                        else: status = str(val)
                                    else:
                                        status = str(val)
                                    pd["settings"][cat_name].append({"name": s.get("name"), "status": status})
                            raw_s = p.get("settings", {})
                            if isinstance(raw_s, dict) and raw_s:
                                sensor_rows = []
                                # Policy meta bilgileri
                                sensor_rows.append({"name": "Policy ID", "status": p.get("id","—")})
                                sensor_rows.append({"name": "Status", "status": "ON" if p.get("enabled") else "OFF"})
                                sensor_rows.append({"name": "Created", "status": p.get("created_timestamp","")[:16].replace("T"," ")})
                                sensor_rows.append({"name": "Last Modified", "status": p.get("modified_timestamp","")[:16].replace("T"," ")})
                                # Sensor version
                                if "sensor_version" in raw_s:
                                    sensor_rows.append({"name": "Sensor Version", "status": f"{raw_s.get('sensor_version','')}  (Build: {raw_s.get('build','')})"})
                                # Stage
                                if "stage" in raw_s:
                                    stage_map = {"prod": "Production (N-1)", "n-1": "N-1", "lt": "Long Term", "early_adopter": "Early Adopter"}
                                    sensor_rows.append({"name": "Update Stage", "status": stage_map.get(raw_s.get("stage",""), raw_s.get("stage",""))})
                                # Early adopter
                                if "show_early_adopter_builds" in raw_s:
                                    sensor_rows.append({"name": "Show Early Adopter Sensor Builds", "status": "ON" if raw_s.get("show_early_adopter_builds") else "OFF"})
                                # LTS/LTV
                                if "is_lts_build" in raw_s:
                                    sensor_rows.append({"name": "Show LTV Sensor Builds", "status": "ON" if raw_s.get("is_lts_build") else "OFF"})
                                # Uninstall protection
                                if "uninstall_protection" in raw_s:
                                    sensor_rows.append({"name": "Uninstall and Maintenance Protection", "status": "ON" if raw_s.get("uninstall_protection") == "ENABLED" else "OFF"})
                                # Scheduler
                                scheduler = raw_s.get("scheduler", {})
                                if isinstance(scheduler, dict):
                                    if scheduler.get("enabled"):
                                        tz = scheduler.get("timezone", "")
                                        schedules = scheduler.get("schedules", [])
                                        sched_str = ""
                                        if schedules:
                                            days_map = {0:"Sun",1:"Mon",2:"Tue",3:"Wed",4:"Thu",5:"Fri",6:"Sat"}
                                            s = schedules[0]
                                            days = ", ".join([days_map.get(d,str(d)) for d in s.get("days",[])])
                                            sched_str = f"{s.get('start','')} - {s.get('end','')}  ({days})  {tz}"
                                        sensor_rows.append({"name": "Time Blocks", "status": sched_str or "Enabled"})
                                    else:
                                        sensor_rows.append({"name": "Time Blocks", "status": "OFF"})
                                if sensor_rows:
                                    pd["settings"]["Sensor Update Settings"] = sensor_rows
                            policy_data[gname][ptype].append(pd)

    # ── EXCLUSIONS (FALCONPY) ──────────────────────────────
    print("  Exclusions çekiliyor...")
    ml_client = MLExclusions(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    r = ml_client.query_exclusions(limit=500)
    ml_ids = r.get("body", {}).get("resources", []) if r['status_code'] == 200 else []
    ml_details = []
    for i in range(0, len(ml_ids), 100):
        r = ml_client.get_exclusions(ids=ml_ids[i:i+100])
        if r['status_code'] == 200:
            ml_details.extend(r.get("body", {}).get("resources", []))

    ioa_client = IOAExclusions(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    r = ioa_client.query_exclusions(limit=500)
    ioa_ids = r.get("body", {}).get("resources", []) if r['status_code'] == 200 else []
    ioa_details = []
    for i in range(0, len(ioa_ids), 100):
        r = ioa_client.get_exclusions(ids=ioa_ids[i:i+100])
        if r['status_code'] == 200:
            ioa_details.extend(r.get("body", {}).get("resources", []))

    sv_client = SensorVisibilityExclusions(client_id=CLIENT_ID, client_secret=SECRET, base_url=BASE_URL)
    r = sv_client.query_exclusions(limit=500)
    sv_ids = r.get("body", {}).get("resources", []) if r['status_code'] == 200 else []
    sv_details = []
    if sv_ids:
        for i in range(0, len(sv_ids), 100):
            r = sv_client.get_exclusions(ids=sv_ids[i:i+100])
            if r['status_code'] == 200:
                sv_details.extend(r.get("body", {}).get("resources", []))

    # ── WORKBOOK OLUŞTUR ────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    group_types = {}
    for g in group_list:
        group_types[g.get("name","")] = g.get("group_type","static").capitalize()

    # ── COVER ───────────────────────────────────────────────
    ws_cover = wb.create_sheet("Cover")
    ws_cover.sheet_properties.tabColor = NAVY
    ws_cover.sheet_view.showGridLines = False
    for row in range(1, 50):
        for col in range(1, 15):
            ws_cover.cell(row=row, column=col).fill = fill(WHITE)
    for col in range(1, 15):
        for r in range(1, 6):
            ws_cover.cell(row=r, column=col).fill = fill(NAVY)
    ws_cover.merge_cells("B2:N4")
    ws_cover["B2"] = "CrowdStrike Falcon"
    ws_cover["B2"].font = Font(color=WHITE, bold=True, size=28, name=FONT)
    ws_cover["B2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 15):
        ws_cover.cell(row=6, column=col).fill = fill(ORANGE)
    ws_cover.row_dimensions[6].height = 6
    ws_cover.merge_cells("B8:N9")
    ws_cover["B8"] = "CarrefourSA"
    ws_cover["B8"].font = Font(color=NAVY, bold=True, size=32, name=FONT)
    ws_cover["B8"].alignment = Alignment(horizontal="left", vertical="center")
    ws_cover.merge_cells("B10:N11")
    ws_cover["B10"] = "Security Health Check Report"
    ws_cover["B10"].font = Font(color=ORANGE, bold=True, size=20, name=FONT)
    ws_cover["B10"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 14):
        ws_cover.cell(row=12, column=col).fill = fill(ORANGE)
    ws_cover.row_dimensions[12].height = 3
    for i, (label, value) in enumerate([
        ("Report Date", today_display), ("Report Period", report_month),
        ("Prepared by", "Pure7"), ("Customer", "CarrefourSA"),
        ("Product", "CrowdStrike Falcon"), ("Version", "1.0"),
    ]):
        r = 14 + i * 2
        ws_cover.merge_cells(f"B{r}:D{r}")
        ws_cover[f"B{r}"] = label
        ws_cover[f"B{r}"].font = fnt(NAVY, bold=True, size=11)
        ws_cover[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws_cover.merge_cells(f"E{r}:N{r}")
        ws_cover[f"E{r}"] = value
        ws_cover[f"E{r}"].font = fnt("333333", size=11)
        ws_cover[f"E{r}"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 15):
        for r in range(45, 50):
            ws_cover.cell(row=r, column=col).fill = fill(NAVY)
    ws_cover.merge_cells("B46:N48")
    ws_cover["B46"] = "Pure7  |  Confidential"
    ws_cover["B46"].font = Font(color=WHITE, size=10, name=FONT)
    ws_cover["B46"].alignment = Alignment(horizontal="center", vertical="center")
    for r in range(1, 50): ws_cover.row_dimensions[r].height = 20
    for r in range(2, 5): ws_cover.row_dimensions[r].height = 28
    ws_cover.row_dimensions[8].height = 40; ws_cover.row_dimensions[10].height = 30
    set_col_width(ws_cover, 1, 3)
    for c in range(2, 15): set_col_width(ws_cover, c, 12)

    # ── SUMMARY ─────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = ORANGE
    ws_sum.sheet_view.showGridLines = False
    ws_sum.merge_cells("A1:L1")
    ws_sum["A1"] = "CarrefourSA – CrowdStrike Security Dashboard"
    ws_sum["A1"].fill = fill(NAVY); ws_sum["A1"].font = fnt(WHITE, bold=True, size=16)
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 40
    ws_sum.merge_cells("A2:L2")
    ws_sum["A2"] = f"Report Date: {today_display}"
    ws_sum["A2"].fill = fill(ORANGE); ws_sum["A2"].font = fnt(WHITE, bold=True, size=10)
    ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    def stat_card(ws, row, col, title, value, bg, fg):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        c1 = ws.cell(row=row, column=col, value=title)
        c1.fill = fill(bg); c1.font = fnt(fg, bold=True, size=9)
        c1.alignment = Alignment(horizontal="center", vertical="center"); c1.border = thin_border()
        c2 = ws.cell(row=row+1, column=col, value=value)
        c2.fill = fill(bg); c2.font = Font(color=fg, bold=True, size=22, name=FONT)
        c2.alignment = Alignment(horizontal="center", vertical="center"); c2.border = thin_border()
        ws.row_dimensions[row].height = 18; ws.row_dimensions[row+1].height = 36

    total = len(hosts)
    online = sum(1 for h in hosts if h.get("status") == "normal")
    offline = total - online
    windows_c = sum(1 for h in hosts if "Windows" in str(h.get("platform_name","")))
    linux_c   = sum(1 for h in hosts if "Linux"   in str(h.get("platform_name","")))
    mac_c     = sum(1 for h in hosts if "Mac"     in str(h.get("platform_name","")))

    ws_sum.merge_cells("A4:L4")
    ws_sum["A4"] = "  HOST INVENTORY"
    ws_sum["A4"].fill = fill(LIGHT_NAVY); ws_sum["A4"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[4].height = 22
    stat_card(ws_sum, 5, 1,  "Total Hosts", total,   LIGHT_NAVY,  NAVY)
    stat_card(ws_sum, 5, 3,  "🟢 Online",   online,  GREEN_BG,    GREEN_FG)
    stat_card(ws_sum, 5, 5,  "🔴 Offline",  offline, RED_BG,      RED_FG)
    stat_card(ws_sum, 5, 7,  "Windows",     windows_c, LIGHT_ORANGE, ORANGE)
    stat_card(ws_sum, 5, 9,  "Linux",       linux_c, LIGHT_NAVY,  NAVY)
    stat_card(ws_sum, 5, 11, "macOS",       mac_c,   GRAY,        "555555")

    grp_total = len(policy_data)
    grp_with  = sum(1 for p in policy_data.values() if p)
    grp_without = grp_total - grp_with
    all_settings = []
    for ptypes in policy_data.values():
        if "Prevention" in ptypes:
            for p in ptypes["Prevention"]:
                for cat in p.get("settings", {}).values():
                    for s in cat:
                        all_settings.append(s.get("status",""))
    on_c  = all_settings.count("ON")
    off_c = all_settings.count("OFF")
    mix_c = len(all_settings) - on_c - off_c

    ws_sum.merge_cells("A8:L8")
    ws_sum["A8"] = "  HOST GROUP & POLICY STATUS"
    ws_sum["A8"].fill = fill(LIGHT_NAVY); ws_sum["A8"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A8"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[8].height = 22
    stat_card(ws_sum, 9, 1,  "Total Groups",   grp_total,   LIGHT_NAVY,  NAVY)
    stat_card(ws_sum, 9, 3,  "✅ Has Policy",   grp_with,    GREEN_BG,    GREEN_FG)
    stat_card(ws_sum, 9, 5,  "⚠️ No Policy",    grp_without, RED_BG,      RED_FG)
    stat_card(ws_sum, 9, 7,  "Settings ON",    on_c,        GREEN_BG,    GREEN_FG)
    stat_card(ws_sum, 9, 9,  "Settings OFF",   off_c,       GRAY,        "808080")
    stat_card(ws_sum, 9, 11, "Mixed (ML/Det)", mix_c,       YELLOW_BG,   YELLOW_FG)

    excl_total = len(ml_details) + len(ioa_details) + len(sv_details)
    ws_sum.merge_cells("A12:L12")
    ws_sum["A12"] = "  EXCLUSIONS"
    ws_sum["A12"].fill = fill(LIGHT_NAVY); ws_sum["A12"].font = fnt(NAVY, bold=True, size=11)
    ws_sum["A12"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[12].height = 22
    stat_card(ws_sum, 13, 1, "ML Exclusions",  len(ml_details),  LIGHT_ORANGE, ORANGE)
    stat_card(ws_sum, 13, 3, "IOA Exclusions", len(ioa_details), LIGHT_NAVY,   NAVY)
    stat_card(ws_sum, 13, 5, "SV Exclusions",  len(sv_details),  GRAY,         "808080")
    stat_card(ws_sum, 13, 7, "Total",          excl_total,       LIGHT_NAVY,   NAVY)
    for c in range(1, 13): set_col_width(ws_sum, c, 10)

    # ── RISK & ACTIONS ──────────────────────────────────────
    ws_risk = wb.create_sheet("Risk & Actions")
    ws_risk.sheet_properties.tabColor = "C00000"
    ws_risk.merge_cells("A1:E1")
    ws_risk["A1"] = "CarrefourSA – Risk & Action Required"
    ws_risk["A1"].fill = fill(NAVY); ws_risk["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_risk["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_risk.row_dimensions[1].height = 35
    ws_risk.merge_cells("A2:E2")
    ws_risk["A2"] = f"Generated: {today_display}  |  Automated risk detection"
    ws_risk["A2"].fill = fill(ORANGE); ws_risk["A2"].font = fnt(WHITE, bold=True, size=9)
    ws_risk["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_risk.row_dimensions[2].height = 18
    for col_i, h in enumerate(["Severity","Category","Issue","Details","Recommendation"],1):
        c = ws_risk.cell(row=3, column=col_i, value=h)
        c.fill = fill(LIGHT_NAVY); c.font = fnt(NAVY, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
    ws_risk.row_dimensions[3].height = 18

    risks = []
    for gname, ptypes in policy_data.items():
        if not ptypes:
            risks.append(("🔴 HIGH","Policy",f"No policy assigned to: {gname}",gname,"Assign at least a Prevention policy immediately."))
    disabled_critical = {
        "Volume Shadow Copy - Protect": "Ransomware protection",
        "Quarantine & Security Center Registration": "Quarantine capability",
        "Quarantine on Write": "Real-time quarantine",
        "Driver Load Prevention": "Kernel-level protection",
        "Credential Dumping": "Credential theft prevention",
    }
    for gname, ptypes in policy_data.items():
        if "Prevention" in ptypes:
            for p in ptypes["Prevention"]:
                for cat, settings in p.get("settings",{}).items():
                    for s in settings:
                        if s.get("name","") in disabled_critical and s.get("status","") == "OFF":
                            risks.append(("🔴 HIGH","Prevention Policy",
                                f"{s['name']} is OFF",
                                f"Policy: {p.get('name','')} | Group: {gname}",
                                disabled_critical[s['name']] + " is disabled."))
    if offline > 0:
        risks.append(("🟡 MEDIUM","Host Availability",f"{offline} hosts offline",
            "Offline/degraded hosts detected",
            "Investigate offline hosts — they may not receive policy updates."))
    ml_dis = []
    for gname, ptypes in policy_data.items():
        if "Prevention" in ptypes:
            for p in ptypes["Prevention"]:
                for cat, settings in p.get("settings",{}).items():
                    for s in settings:
                        raw = s.get("status","")
                        try:
                            val = ast.literal_eval(raw) if "{" in str(raw) else {}
                            if isinstance(val, dict) and val.get("prevention","") == "DISABLED":
                                ml_dis.append(s.get("name",""))
                        except: pass
    if ml_dis:
        risks.append(("🟡 MEDIUM","ML Prevention",f"{len(ml_dis)} ML settings Prevention DISABLED",
            "; ".join(ml_dis[:3]) + ("..." if len(ml_dis)>3 else ""),
            "Consider enabling Prevention mode for ML detections."))

    sev_colors = {"🔴 HIGH":(RED_FG,RED_BG),"🟡 MEDIUM":(YELLOW_FG,YELLOW_BG),"🟢 LOW":(GREEN_FG,GREEN_BG)}
    for idx, (sev, cat, issue, details, rec) in enumerate(risks):
        row = idx + 4
        fg, bg = sev_colors.get(sev, ("000000", WHITE))
        for col_i, val in enumerate([sev, cat, issue, details, rec], 1):
            c = ws_risk.cell(row=row, column=col_i, value=val)
            c.fill = fill(bg); c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.font = fnt(fg, bold=True, size=9) if col_i==1 else (fnt("444444",size=8,italic=True) if col_i==5 else fnt(size=9))
        ws_risk.row_dimensions[row].height = 30
    set_col_width(ws_risk,1,14); set_col_width(ws_risk,2,22)
    set_col_width(ws_risk,3,38); set_col_width(ws_risk,4,45); set_col_width(ws_risk,5,55)

    # ── ALL HOSTS ───────────────────────────────────────────
    print("  All Hosts sheet yazılıyor...")
    ws_hosts = wb.create_sheet("All Hosts")
    ws_hosts.sheet_properties.tabColor = NAVY
    HEADERS = ["Hostname","Status","Platform","Device Type","OS Version","Local IP",
               "Sensor Version","Last Seen","Host Groups","Prevention Policy","Sensor Update Policy","Last Logged In User"]
    ws_hosts.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws_hosts["A1"] = f"CarrefourSA – CrowdStrike Host Inventory  ({len(hosts)} hosts)"
    ws_hosts["A1"].fill = fill(NAVY); ws_hosts["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_hosts["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_hosts.row_dimensions[1].height = 35
    for col_i, h in enumerate(HEADERS, 1):
        c = ws_hosts.cell(row=2, column=col_i, value=h)
        c.fill = fill(ORANGE); c.font = fnt(WHITE, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = thin_border()
    ws_hosts.row_dimensions[2].height = 22
    for idx, host in enumerate(sorted(hosts, key=lambda x: x.get("hostname","").lower()), 1):
        row = idx + 2
        bg = WHITE if idx % 2 != 0 else GRAY
        status = host.get("status","unknown")
        if status=="normal": sd="🟢 Online"; rbg=bg
        elif status=="contained": sd="🔴 Contained"; rbg=RED_BG
        elif status in ("reduced_functionality","degraded"): sd="🟡 Degraded"; rbg=YELLOW_BG
        else: sd=f"⚫ {status.replace('_',' ').title()}"; rbg=bg
        hgids = host.get("groups",[])
        hgnames = ", ".join([groups.get(gid,gid[:8]) for gid in hgids]) if hgids else "—"
        dp = host.get("device_policies",{})
        prev_pol = "—"; sensor_pol = "—"
        if isinstance(dp,dict):
            pi = dp.get("prevention",{})
            if pi:
                pid = pi.get("policy_id","")
                prev_pol = policies.get(pid,{}).get("name", pi.get("applied_status","—")) or "—"
            si = dp.get("sensor_update",{})
            if si:
                sid = si.get("policy_id","")
                sensor_pol = policies.get(sid,{}).get("name", si.get("applied_status","—")) or "—"
        row_data = [host.get("hostname","—"), sd, host.get("platform_name","—"),
                    host.get("product_type_desc","—"), host.get("os_version","—"),
                    host.get("local_ip","—"), host.get("agent_version","—"),
                    fmt_date(host.get("last_seen","")), hgnames, prev_pol, sensor_pol,
                    host.get("last_login_user","—")]
        for col_i, val in enumerate(row_data, 1):
            c = ws_hosts.cell(row=row, column=col_i, value=val)
            c.fill = fill(rbg); c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")
            if col_i==1: c.font = fnt(NAVY, bold=True, size=9)
            elif col_i==2:
                if "Online" in str(val): c.font = fnt(GREEN_FG, bold=True, size=9)
                elif "Contained" in str(val): c.font = fnt(RED_FG, bold=True, size=9)
                elif "Degraded" in str(val): c.font = fnt(YELLOW_FG, bold=True, size=9)
                else: c.font = fnt(size=9)
            else: c.font = fnt(size=9)
        ws_hosts.row_dimensions[row].height = 16
    ws_hosts.freeze_panes = "A3"
    for col_i, w in enumerate([22,14,12,18,22,14,14,18,38,38,32,22],1):
        set_col_width(ws_hosts, col_i, w)

    # ── HOST GRUPLARI ───────────────────────────────────────
    ws1 = wb.create_sheet("Host Grupları"); set_outline(ws1)
    ws1.sheet_properties.tabColor = ORANGE
    ws1.merge_cells("A1:E1")
    ws1["A1"] = "CarrefourSA – CrowdStrike Host Grupları"
    ws1["A1"].fill = fill(NAVY); ws1["A1"].font = fnt(WHITE, bold=True, size=14)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35
    row = 2
    for gtype_label in ["STATIC GRUPLAR","DYNAMIC GRUPLAR"]:
        is_dyn = "DYNAMIC" in gtype_label
        filtered = [(g, p) for g, p in policy_data.items()
                    if group_types.get(g,"Static") == ("Dynamic" if is_dyn else "Static")]
        ws1.merge_cells(f"A{row}:E{row}")
        ws1[f"A{row}"] = f"  ▶  {gtype_label}"
        ws1[f"A{row}"].fill = fill(NAVY); ws1[f"A{row}"].font = fnt(WHITE, bold=True, size=12)
        ws1[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws1.row_dimensions[row].height = 26; row += 1
        for col_i, h in enumerate(["#","Host Grubu","Tip","Policy Sayısı","Atanan Policy Tipleri"],1):
            c = ws1.cell(row=row, column=col_i, value=h)
            c.fill = fill(LIGHT_NAVY); c.font = fnt(NAVY, bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
        hide_row(ws1, row, 1); ws1.row_dimensions[row].height = 16; row += 1
        for i, (gname, ptypes) in enumerate(filtered, 1):
            count = sum(len(v) for v in ptypes.values())
            ptype_list = ", ".join(ptypes.keys()) if ptypes else "⚠️ Yok"
            bg = LIGHT_ORANGE if i%2==0 else WHITE
            for col_i, val in enumerate([i, gname, group_types.get(gname,""), count, ptype_list], 1):
                c = ws1.cell(row=row, column=col_i, value=val)
                c.fill = fill(bg); c.border = thin_border()
                c.alignment = Alignment(horizontal="center" if col_i in (1,3,4) else "left", vertical="center", wrap_text=True)
                c.font = fnt(RED_FG, bold=True, size=9) if col_i==4 and count==0 else fnt(size=9)
            hide_row(ws1, row, 1); ws1.row_dimensions[row].height = 18; row += 1
        ws1.merge_cells(f"A{row}:E{row}"); ws1[f"A{row}"] = ""
        ws1.row_dimensions[row].height = 10; row += 1
    set_col_width(ws1,1,5); set_col_width(ws1,2,28); set_col_width(ws1,3,12)
    set_col_width(ws1,4,16); set_col_width(ws1,5,55)

    # ── POLICY ATAMALARI ────────────────────────────────────
    ws2 = wb.create_sheet("Policy Atamaları"); set_outline(ws2)
    ws2.sheet_properties.tabColor = ORANGE
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "CarrefourSA – Host Grubu / Policy Atamaları"
    ws2["A1"].fill = fill(NAVY); ws2["A1"].font = fnt(WHITE, bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    row = 2
    for gname, ptypes in policy_data.items():
        ws2.merge_cells(f"A{row}:D{row}")
        c = ws2[f"A{row}"]
        c.value = f"  ▶  📁  {gname}"
        c.fill = fill(NAVY); c.font = fnt(WHITE, bold=True, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row].height = 26
        t = Side(style='medium', color=ORANGE)
        for col_i in range(1,5):
            ws2.cell(row=row, column=col_i).border = Border(left=t,right=t,top=t,bottom=Side(style='thin',color="CCCCCC"))
        row += 1
        if not ptypes:
            ws2.merge_cells(f"A{row}:D{row}")
            c = ws2[f"A{row}"]
            c.value = "     ⚠️  Bu host grubuna policy atanmamış"
            c.fill = fill(YELLOW_BG); c.font = fnt(YELLOW_FG, bold=True, size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            t = Side(style='medium', color=ORANGE); th = Side(style='thin', color="CCCCCC")
            for col_i in range(1,5):
                ws2.cell(row=row,column=col_i).border = Border(left=t if col_i==1 else th,right=t if col_i==4 else th,top=th,bottom=t)
            hide_row(ws2, row, 1); ws2.row_dimensions[row].height = 20; row += 1
        else:
            for col_i, h in enumerate(["Policy Tipi","Policy Adı","Platform","Durum"],1):
                c = ws2.cell(row=row, column=col_i, value=h)
                c.fill = fill(LIGHT_NAVY); c.font = fnt(NAVY, bold=True, size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                t = Side(style='medium', color=ORANGE); th = Side(style='thin', color="CCCCCC")
                c.border = Border(left=t if col_i==1 else th, right=t if col_i==4 else th, top=th, bottom=th)
            hide_row(ws2, row, 1); ws2.row_dimensions[row].height = 16; row += 1
            all_pol = [(pt, p) for pt, ps in ptypes.items() for p in ps]
            for idx2, (ptype, p) in enumerate(all_pol):
                is_last = idx2 == len(all_pol)-1
                status = "✅  Aktif" if p.get("enabled") else "❌  Pasif"
                bg = LIGHT_ORANGE if idx2%2==0 else WHITE
                t = Side(style='medium', color=ORANGE); th = Side(style='thin', color="CCCCCC")
                for col_i, val in enumerate([ptype, p.get("name",""), p.get("platform",""), status],1):
                    c = ws2.cell(row=row, column=col_i, value=val)
                    c.fill = fill(bg)
                    c.alignment = Alignment(horizontal="center" if col_i in (1,3,4) else "left", vertical="center")
                    c.border = Border(left=t if col_i==1 else th, right=t if col_i==4 else th, top=th, bottom=t if is_last else th)
                    if col_i==1: c.font = fnt(ORANGE, bold=True, size=9)
                    elif col_i==4: c.font = fnt(GREEN_FG if "Aktif" in status else RED_FG, bold=True, size=9)
                    else: c.font = fnt(size=9)
                hide_row(ws2, row, 1); ws2.row_dimensions[row].height = 18; row += 1
        ws2.merge_cells(f"A{row}:D{row}"); ws2[f"A{row}"] = ""
        ws2.row_dimensions[row].height = 10; row += 1
    set_col_width(ws2,1,22); set_col_width(ws2,2,48)
    set_col_width(ws2,3,14); set_col_width(ws2,4,14)

    # ── POLICY DETAYLARI ────────────────────────────────────
    ws3 = wb.create_sheet("Policy Detayları"); set_outline(ws3)
    ws3.sheet_properties.tabColor = ORANGE
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "CarrefourSA – CrowdStrike Policy Details"
    ws3["A1"].fill = fill(NAVY); ws3["A1"].font = fnt(WHITE, bold=True, size=14)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35
    POLICY_TYPES = ["Prevention","Sensor Update","Firewall","Device Control","Response"]
    row = 2
    for ptype in POLICY_TYPES:
        policies_found = {}
        for gname, ptypes in policy_data.items():
            if ptype in ptypes:
                for p in ptypes[ptype]:
                    pname = p.get("name","")
                    if pname not in policies_found:
                        policies_found[pname] = {"policy": p, "groups": []}
                    policies_found[pname]["groups"].append(gname)
        ws3.merge_cells(f"A{row}:D{row}")
        ws3[f"A{row}"] = f"  ▶  {ptype.upper()} POLICIES"
        ws3[f"A{row}"].fill = fill(NAVY); ws3[f"A{row}"].font = fnt(WHITE, bold=True, size=12)
        ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[row].height = 28; row += 1
        if not policies_found:
            ws3.merge_cells(f"A{row}:D{row}")
            ws3[f"A{row}"] = f"     ⚠️  No {ptype} policies assigned"
            ws3[f"A{row}"].fill = fill(YELLOW_BG); ws3[f"A{row}"].font = fnt(YELLOW_FG, bold=True, size=10)
            ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            hide_row(ws3, row, 1); ws3.row_dimensions[row].height = 20; row += 2; continue
        for pname, pinfo in policies_found.items():
            p = pinfo["policy"]; groups_str = ", ".join(pinfo["groups"])
            ws3.merge_cells(f"A{row}:D{row}")
            ws3[f"A{row}"] = f"    ▶  {pname}"
            ws3[f"A{row}"].fill = fill(ORANGE); ws3[f"A{row}"].font = fnt(WHITE, bold=True, size=11)
            ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws3.row_dimensions[row].height = 24; ws3.row_dimensions[row].outline_level = 1; row += 1
            ws3.merge_cells(f"A{row}:D{row}")
            ws3[f"A{row}"] = f"       🔗 Platform: {p.get('platform','')}  |  {'✅ Active' if p.get('enabled') else '❌ Inactive'}  |  Groups: {groups_str}"
            ws3[f"A{row}"].fill = fill(LIGHT_ORANGE); ws3[f"A{row}"].font = fnt(NAVY, italic=True, size=9)
            ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            hide_row(ws3, row, 2); ws3.row_dimensions[row].height = 16; row += 1
            for col_i, h in enumerate(["Category","Setting","Status","Description"],1):
                c = ws3.cell(row=row, column=col_i, value=h)
                c.fill = fill(LIGHT_NAVY); c.font = fnt(NAVY, bold=True, size=9)
                c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
            hide_row(ws3, row, 2); ws3.row_dimensions[row].height = 16; row += 1
            settings = p.get("settings",{})
            if not settings:
                ws3.merge_cells(f"A{row}:D{row}")
                ws3[f"A{row}"] = "     No detail settings available"
                ws3[f"A{row}"].fill = fill(GRAY); ws3[f"A{row}"].font = fnt("808080", italic=True, size=9)
                hide_row(ws3, row, 2); ws3.row_dimensions[row].height = 15; row += 1
            else:
                for cat_name, cat_settings in settings.items():
                    first_in_cat = True
                    for s in cat_settings:
                        raw_status = s.get("status",""); sname = s.get("name","")
                        desc = DESCRIPTIONS.get(sname, "—")
                        stype, det, prev = parse_status(raw_status)
                        if stype == "mixed":
                            fg_s, bg_s = level_style(det)
                            display = f"Detection:  {det}\nPrevention: {prev}"; rh=32
                        elif det == "ON":
                            fg_s, bg_s = level_style("ON")
                            display = "✅  ON"; rh=22
                        elif det == "OFF":
                            fg_s, bg_s = level_style("OFF")
                            display = "❌  OFF"; rh=22
                        else:
                            # Düz değer: UUID, tarih, versiyon, vb.
                            fg_s = NAVY; bg_s = LIGHT_NAVY
                            display = str(det); rh=22
                        ca = ws3.cell(row=row, column=1, value=cat_name if first_in_cat else "")
                        ca.fill = fill(bg_s); ca.font = fnt(NAVY, bold=True, size=9)
                        ca.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); ca.border = thin_border()
                        cb = ws3.cell(row=row, column=2, value=sname)
                        cb.fill = fill(bg_s); cb.font = fnt(size=9)
                        cb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); cb.border = thin_border()
                        cc = ws3.cell(row=row, column=3, value=display)
                        cc.fill = fill(bg_s); cc.font = fnt(fg_s, bold=True, size=9)
                        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cc.border = thin_border()
                        cd = ws3.cell(row=row, column=4, value=desc)
                        cd.fill = fill(bg_s); cd.font = fnt("444444", size=8, italic=True)
                        cd.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); cd.border = thin_border()
                        hide_row(ws3, row, 2); ws3.row_dimensions[row].height = rh
                        first_in_cat = False; row += 1
            ws3.merge_cells(f"A{row}:D{row}"); ws3[f"A{row}"] = ""
            hide_row(ws3, row, 1); ws3.row_dimensions[row].height = 6; row += 1
        ws3.merge_cells(f"A{row}:D{row}"); ws3[f"A{row}"] = ""
        ws3.row_dimensions[row].height = 10; row += 1
    set_col_width(ws3,1,35); set_col_width(ws3,2,45)
    set_col_width(ws3,3,22); set_col_width(ws3,4,55)

    # ── EXCLUSIONS ──────────────────────────────────────────
    ws_excl = wb.create_sheet("Exclusions")
    ws_excl.sheet_properties.tabColor = "808080"
    EXCL_HEADERS = ["Type","Value","Comment","Applied Globally","Groups","Created By","Created On","Modified By","Modified On"]
    ws_excl.merge_cells(f"A1:{get_column_letter(len(EXCL_HEADERS))}1")
    ws_excl["A1"] = f"CarrefourSA – CrowdStrike Exclusions  ({len(ml_details)+len(ioa_details)+len(sv_details)} total)"
    ws_excl["A1"].fill = fill(NAVY); ws_excl["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_excl["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_excl.row_dimensions[1].height = 35
    for col_i, h in enumerate(EXCL_HEADERS,1):
        c = ws_excl.cell(row=2, column=col_i, value=h)
        c.fill = fill(ORANGE); c.font = fnt(WHITE, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
    ws_excl.row_dimensions[2].height = 20
    erow = 3
    for excl_list, excl_type in [(ml_details,"ML Exclusion"),(ioa_details,"IOA Exclusion"),(sv_details,"SV Exclusion")]:
        if not excl_list: continue
        ws_excl.merge_cells(f"A{erow}:{get_column_letter(len(EXCL_HEADERS))}{erow}")
        ws_excl[f"A{erow}"] = f"  ▶  {excl_type.upper()}S  ({len(excl_list)} items)"
        ws_excl[f"A{erow}"].fill = fill(NAVY); ws_excl[f"A{erow}"].font = fnt(WHITE, bold=True, size=11)
        ws_excl[f"A{erow}"].alignment = Alignment(horizontal="left", vertical="center")
        ws_excl.row_dimensions[erow].height = 22; erow += 1
        for idx, e in enumerate(excl_list):
            bg = WHITE if idx%2==0 else GRAY
            groups_list = e.get("groups",[])
            if isinstance(groups_list, list):
                gs = ", ".join([g.get("name",g.get("id","")) if isinstance(g,dict) else str(g) for g in groups_list])
            else: gs = str(groups_list)
            if not gs: gs = "All Groups" if e.get("applied_globally") else "—"
            applied = "✅ Yes" if e.get("applied_globally") else "❌ No"
            row_data = [excl_type, e.get("value",e.get("pattern_name","—")),
                        e.get("comment","—") or "—", applied, gs,
                        e.get("created_by","—"), fmt_date(e.get("created_timestamp","")),
                        e.get("modified_by","—"), fmt_date(e.get("modified_timestamp",""))]
            for col_i, val in enumerate(row_data,1):
                c = ws_excl.cell(row=erow, column=col_i, value=val)
                c.fill = fill(bg); c.border = thin_border()
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if col_i==1: c.font = fnt(ORANGE, bold=True, size=9)
                elif col_i==2: c.font = fnt(NAVY, size=9)
                elif col_i==4:
                    c.font = fnt(GREEN_FG, bold=True, size=9) if "Yes" in str(val) else fnt(RED_FG, bold=True, size=9)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else: c.font = fnt(size=9)
            ws_excl.row_dimensions[erow].height = 18; erow += 1
        ws_excl.merge_cells(f"A{erow}:{get_column_letter(len(EXCL_HEADERS))}{erow}")
        ws_excl[f"A{erow}"] = ""; ws_excl.row_dimensions[erow].height = 8; erow += 1
    ws_excl.freeze_panes = "A3"
    for col_i, w in enumerate([16,55,30,16,30,25,16,25,16],1):
        set_col_width(ws_excl, col_i, w)

    # ── UNMANAGED ASSETS (FALCONPY) ─────────────────────────
    print("  Unmanaged Assets çekiliyor...")
    import ipaddress as ipmod

    ua_all_ids = []
    ua_offset = 0
    while True:
        r_ua = discover_client.query_hosts(
            filter="entity_type:'unmanaged'+data_providers:'Falcon passive discovery'",
            sort="last_seen_timestamp.desc",
            limit=100,
            offset=ua_offset
        )
        if r_ua['status_code'] != 200:
            print(f"  ⚠️  Error querying unmanaged assets: {r_ua.get('body',{}).get('errors', [])}")
            break
        ua_data = r_ua.get("body", {})
        ua_ids = ua_data.get("resources", [])
        if not ua_ids:
            break
        ua_all_ids.extend(ua_ids)
        ua_offset += len(ua_ids)
        total = ua_data.get("meta", {}).get("pagination", {}).get("total", 0)
        if ua_offset >= total:
            break

    ua_hosts_raw = []
    for i in range(0, len(ua_all_ids), 100):
        batch = ua_all_ids[i:i+100]
        r_ua2 = discover_client.get_hosts(ids=batch)
        if r_ua2['status_code'] == 200:
            ua_hosts_raw.extend(r_ua2.get("body", {}).get("resources", []))

    ua_filtered = []
    for h in ua_hosts_raw:
        ips = [ni.get("local_ip","") for ni in (h.get("network_interfaces") or [])]
        if not ips:
            ip = h.get("current_local_ip","")
            ips = [ip] if ip else []
        has_10x = any(str(ipmod.ip_address(ip)).startswith("10.") for ip in ips if ip)
        if not has_10x: continue
        discoverers = h.get("discoverer_hostnames") or []
        clean = [d for d in discoverers if "$" not in d]
        if not clean: continue
        h["_ips"] = ips
        h["_clean_discoverers"] = clean
        ua_filtered.append(h)

    print(f"  Unmanaged: {len(ua_filtered)} host")

    ws_ua = wb.create_sheet("Unmanaged Assets")
    ws_ua.sheet_properties.tabColor = "7030A0"

    UA_HEADERS = ["Hostname (Discoverer)", "IP Address History", "Manufacturer",
                  "Data Providers", "Entity Type", "Confidence",
                  "First Seen", "Last Seen", "Review Status"]

    ws_ua.merge_cells(f"A1:{get_column_letter(len(UA_HEADERS))}1")
    ws_ua["A1"] = f"CarrefourSA – Unmanaged Assets  ({len(ua_filtered)} assets)"
    ws_ua["A1"].fill = fill(NAVY); ws_ua["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_ua["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_ua.row_dimensions[1].height = 35

    for col_i, h in enumerate(UA_HEADERS, 1):
        c = ws_ua.cell(row=2, column=col_i, value=h)
        c.fill = fill(ORANGE); c.font = fnt(WHITE, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws_ua.row_dimensions[2].height = 22

    conf_map = {0:"Low", 25:"Low", 50:"Medium", 75:"High", 100:"Very High"}

    row = 3
    for idx, h in enumerate(sorted(ua_filtered, key=lambda x: x.get("last_seen_timestamp",""), reverse=True), 1):
        bg = WHITE if idx % 2 != 0 else GRAY

        conf_val = h.get("confidence", 0)
        conf_label = conf_map.get(conf_val, str(conf_val))
        if conf_val >= 75: conf_color = GREEN_FG; conf_bg = GREEN_BG
        elif conf_val >= 50: conf_color = YELLOW_FG; conf_bg = YELLOW_BG
        else: conf_color = RED_FG; conf_bg = RED_BG

        # Her discoverer için ayrı satır
        discoverers = h["_clean_discoverers"]
        ips_str = ", ".join(h["_ips"])
        manufacturer = h.get("system_manufacturer","—") or "—"
        data_providers = ", ".join(h.get("data_providers") or [])
        entity_type = h.get("entity_type","—").capitalize()
        first_seen = fmt_date(h.get("first_seen_timestamp",""))
        last_seen = fmt_date(h.get("last_seen_timestamp",""))
        review_status = h.get("review_status","Not Reviewed") or "Not Reviewed"

        for disc_idx, discoverer in enumerate(discoverers):
            row_data = [
                discoverer,  # Her satırda tek bir discoverer
                ips_str if disc_idx == 0 else "",  # IP'ler sadece ilk satırda
                manufacturer if disc_idx == 0 else "",
                data_providers if disc_idx == 0 else "",
                entity_type if disc_idx == 0 else "",
                conf_label if disc_idx == 0 else "",
                first_seen if disc_idx == 0 else "",
                last_seen if disc_idx == 0 else "",
                review_status if disc_idx == 0 else "",
            ]

            for col_i, val in enumerate(row_data, 1):
                c = ws_ua.cell(row=row, column=col_i, value=val)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if col_i == 6:
                    c.fill = fill(conf_bg); c.font = fnt(conf_color, bold=True, size=9)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.fill = fill(bg); c.font = fnt(NAVY if col_i==1 else "000000", bold=(col_i==1), size=9)
            ws_ua.row_dimensions[row].height = 18
            row += 1

    ws_ua.freeze_panes = "A3"
    for col_i, w in enumerate([35, 20, 25, 22, 14, 12, 16, 16, 16], 1):
        set_col_width(ws_ua, col_i, w)

    # ── HIGH MEMORY USAGE (FALCONPY) ────────────────────────
    print("  High Memory Usage çekiliyor...")

    hm_all_ids = []
    hm_offset = 0
    while True:
        kwargs = {
            "filter": "average_memory_usage_pct:>80",
            "sort": "average_memory_usage_pct.desc",
            "limit": 100,
            "offset": hm_offset
        }
        r_hm = discover_client.query_hosts(**kwargs)
        if r_hm['status_code'] != 200:
            print(f"  ⚠️  Error querying high memory hosts: {r_hm.get('body',{}).get('errors', [])}")
            break
        hm_data = r_hm.get("body", {})
        hm_ids = hm_data.get("resources", [])
        if not hm_ids:
            break
        hm_all_ids.extend(hm_ids)
        hm_offset += len(hm_ids)
        total = hm_data.get("meta", {}).get("pagination", {}).get("total", 0)
        if hm_offset >= total:
            break

    hm_hosts = []
    for i in range(0, len(hm_all_ids), 100):
        batch = hm_all_ids[i:i+100]
        r_hm2 = discover_client.get_hosts(ids=batch)
        if r_hm2['status_code'] == 200:
            hm_hosts.extend(r_hm2.get("body", {}).get("resources", []))

    hm_hosts.sort(key=lambda x: x.get("average_memory_usage_pct", 0), reverse=True)
    print(f"  High Memory: {len(hm_hosts)} host")

    ws_hm = wb.create_sheet("High Memory Usage")
    ws_hm.sheet_properties.tabColor = RED_FG

    HM_HEADERS = ["Hostname", "OS Version", "Platform",
                  "Total Memory (MB)", "Avg Memory Usage (MB)", "Avg Memory %",
                  "Max Memory %", "Avg CPU %", "Max CPU %",
                  "Total Disk (MB)", "Used Disk %", "Last Seen"]

    ws_hm.merge_cells(f"A1:{get_column_letter(len(HM_HEADERS))}1")
    ws_hm["A1"] = f"CarrefourSA – High Memory Usage (>80%)  ({len(hm_hosts)} hosts)"
    ws_hm["A1"].fill = fill(NAVY); ws_hm["A1"].font = fnt(WHITE, bold=True, size=14)
    ws_hm["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_hm.row_dimensions[1].height = 35

    for col_i, h in enumerate(HM_HEADERS, 1):
        c = ws_hm.cell(row=2, column=col_i, value=h)
        c.fill = fill(ORANGE); c.font = fnt(WHITE, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws_hm.row_dimensions[2].height = 22

    for idx, h in enumerate(hm_hosts, 1):
        row = idx + 2
        mem_pct = h.get("average_memory_usage_pct", 0)
        if mem_pct >= 95: bg = RED_BG
        elif mem_pct >= 90: bg = YELLOW_BG
        else: bg = WHITE if idx % 2 != 0 else GRAY

        row_data = [
            h.get("hostname","—") or "—",
            h.get("os_version","—") or "—",
            h.get("platform_name","—") or "—",
            h.get("total_memory","—"),
            h.get("average_memory_usage","—"),
            f"{mem_pct}%",
            f"{h.get('max_memory_usage_pct','—')}%",
            f"{h.get('average_processor_usage','—')}%",
            f"{h.get('max_processor_usage','—')}%",
            h.get("total_disk_space","—"),
            f"{h.get('used_disk_space_pct','—')}%",
            fmt_date(h.get("last_seen_timestamp","")),
        ]

        for col_i, val in enumerate(row_data, 1):
            c = ws_hm.cell(row=row, column=col_i, value=val)
            c.fill = fill(bg); c.border = thin_border()
            c.alignment = Alignment(horizontal="left" if col_i in (1,2,3) else "center",
                                    vertical="center")
            if col_i == 1: c.font = fnt(NAVY, bold=True, size=9)
            elif col_i == 6:
                if mem_pct >= 95: c.font = fnt(RED_FG, bold=True, size=9)
                elif mem_pct >= 90: c.font = fnt(YELLOW_FG, bold=True, size=9)
                else: c.font = fnt(GREEN_FG, bold=True, size=9)
            else: c.font = fnt(size=9)
        ws_hm.row_dimensions[row].height = 16

    ws_hm.freeze_panes = "A3"
    for col_i, w in enumerate([25, 20, 12, 18, 22, 14, 12, 10, 10, 18, 12, 16], 1):
        set_col_width(ws_hm, col_i, w)

    # ── KAYDET ──────────────────────────────────────────────
    filename = f"CarrefourSA_CrowdStrike_HealthCheck_{date_str}.xlsx"
    output_path = f"{OUTPUT_DIR}/{filename}"
    wb.save(output_path)
    print(f"✅ Rapor kaydedildi: {output_path}")
    return output_path

if __name__ == "__main__":
    main()
